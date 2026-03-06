# scripts/export_builders_excel.py
import sys
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Add parent dir to path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from config import PROJECTS_CLEAN_FIXED_FILE

def export_builders_by_location(output_file="builders_by_state_city.xlsx"):
    print(f"[EXPORT] Loading data from {PROJECTS_CLEAN_FIXED_FILE}...")
    try:
        df = pd.read_csv(PROJECTS_CLEAN_FIXED_FILE)
    except Exception as e:
        print(f"[ERROR] Failed to read CSV: {e}")
        return

    # User requires builder name, project city, project state, and builder link
    # project_city as 'city', project_state as 'state'
    # builder_website as 'builder_link'
    
    required_cols = ['builder_name', 'city', 'state', 'builder_website']
    for col in required_cols:
        if col not in df.columns:
            print(f"[ERROR] Missing required column: {col}")
            return

    # Get unique builders per city/state
    df_builders = df[required_cols].drop_duplicates()
    df_builders = df_builders.rename(columns={'builder_website': 'builder_link'})
    
    # Sort for consistent output
    df_builders = df_builders.sort_values(['state', 'city', 'builder_name'])

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    city_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    center_align = Alignment(horizontal="center")
    border = Border(left=Side(style='thin'), 
                   right=Side(style='thin'), 
                   top=Side(style='thin'), 
                   bottom=Side(style='thin'))

    # Group by state -> Sheet
    states = df_builders['state'].unique()
    for state in states:
        # Sheet name limit is 31 chars, handles 'Unknown' or long names
        sheet_name = str(state)[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        # Filter for this state
        df_state = df_builders[df_builders['state'] == state]
        cities = df_state['city'].unique()
        
        current_row = 1
        
        for city in cities:
            # City Header (Merged Row)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            cell = ws.cell(row=current_row, column=1, value=str(city))
            cell.alignment = center_align
            cell.font = Font(bold=True, size=12)
            cell.fill = city_fill
            current_row += 1
            
            # Column Headers
            headers = ['builder name', 'builder city', 'builder state', 'builder link']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            current_row += 1
            
            # Builder Rows
            df_city = df_state[df_state['city'] == city]
            for _, row in df_city.iterrows():
                ws.cell(row=current_row, column=1, value=row['builder_name']).border = border
                ws.cell(row=current_row, column=2, value=row['city']).border = border
                ws.cell(row=current_row, column=3, value=row['state']).border = border
                ws.cell(row=current_row, column=4, value=row['builder_link']).border = border
                current_row += 1
            
            # Empty row after city block
            current_row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 60

    wb.save(output_file)
    print(f"[OK] Excel report saved to {output_file}")

if __name__ == "__main__":
    export_builders_by_location()
